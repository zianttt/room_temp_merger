from datetime import datetime
import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
import re

def auto_detect_sheets(wb):
    """
    Auto-detect sheets by checking the sheet titles and (if necessary) early cell values.
    Expected roles: 'Max', 'Min', 'Midband', and 'Room Data'
    """
    role_keywords = {
        "Max": ["max", "maximum"],
        "Min": ["min", "minimum"],
        "Midband": ["midband"],
        "Room Data": ["sensed value", "room"]
    }
    detected = {role: None for role in role_keywords}
    
    # First pass: check sheet titles
    for sheet in wb.worksheets:
        sheet_title_lower = sheet.title.lower()
        for role, keywords in role_keywords.items():
            if any(keyword in sheet_title_lower for keyword in keywords):
                if detected[role] is None:
                    detected[role] = sheet.title
                    
    # Second pass: if a role is still not detected, check the first 10x10 cells for keywords.
    for role, value in detected.items():
        if value is None:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = cell.value.lower()
                            if any(keyword in cell_text for keyword in role_keywords[role]):
                                detected[role] = sheet.title
                                break
                    if detected[role] is not None:
                        break
                if detected[role] is not None:
                    break
    return detected

def detect_data_offset(sheet, header_rows_count=3, header_cols_count=2):
    """
    Detect the start of numeric data in the sheet. This is used for copying headers.
    """
    data_start_row = None
    data_start_col = None
    for r in range(sheet.min_row, sheet.max_row + 1):
        for c in range(sheet.min_column, sheet.max_column + 1):
            cell_val = sheet.cell(row=r, column=c).value
            if isinstance(cell_val, (int, float)):
                data_start_row = r
                data_start_col = c
                break
        if data_start_row is not None:
            break
    # Determine header boundaries (if data is found)
    if data_start_row is not None:
        header_row_start = max(sheet.min_row, data_start_row - header_rows_count)
        header_col_start = max(sheet.min_column, data_start_col - header_cols_count)
    else:
        # Fallback: if no numeric data is found, assume the whole sheet is data.
        header_row_start = sheet.min_row
        header_col_start = sheet.min_column
        data_start_row = sheet.min_row
        data_start_col = sheet.min_column
        
    return header_row_start, header_col_start, data_start_row, data_start_col

def build_sensor_mapping(sheet, header_row=2):
    """
    Build a mapping from sensor ID (e.g., 'OC000011') to column index,
    based on the header row where sensor IDs are present.
    """
    mapping = {}
    for col in range(sheet.min_column, sheet.max_column + 1):
        header_val = sheet.cell(row=header_row, column=col).value
        if header_val and isinstance(header_val, str):
            match = re.search(r"OC\d+", header_val)
            if match:
                sensor_id = match.group()
                mapping[sensor_id] = col
    return mapping

def process_excel(file) -> BytesIO:
    wb = openpyxl.load_workbook(file)
    
    # Auto-detect sheets based on keywords.
    detected_mapping = auto_detect_sheets(wb)
    room_sheet_name = detected_mapping.get("Room Data")
    min_sheet_name = detected_mapping.get("Min")
    max_sheet_name = detected_mapping.get("Max")
    mid_sheet_name = detected_mapping.get("Midband")  # Optional
    
    # Verify required sheets.
    missing_sheets = []
    for name, sheet in [("Room Data", room_sheet_name), ("Min", min_sheet_name), ("Max", max_sheet_name)]:
        if sheet is None:
            missing_sheets.append(name)
    if missing_sheets:
        st.error("The following required sheets are missing: " + ", ".join(missing_sheets))
        return None
    if mid_sheet_name is None:
        st.warning("Sheet for 'Midband' not found. It will be ignored.")
    
    room_sheet = wb[room_sheet_name]
    min_sheet = wb[min_sheet_name]
    max_sheet = wb[max_sheet_name]
    
    # Build sensor mappings using the header row where sensor IDs exist (assuming row 2)
    room_mapping = build_sensor_mapping(room_sheet, header_row=2)
    min_mapping  = build_sensor_mapping(min_sheet, header_row=2)
    max_mapping  = build_sensor_mapping(max_sheet, header_row=2)
    
    # Detect the starting row for data in the Room sheet (headers will be copied directly)
    _, _, room_data_start_row, room_data_start_col = detect_data_offset(room_sheet, header_rows_count=3, header_cols_count=2)
    
    # Remove existing "Result" sheet if it exists and create a new one.
    if "Result" in wb.sheetnames:
        wb.remove(wb["Result"])
    result_sheet = wb.create_sheet("Result")
    
    # Define cell fills for formatting.
    blue_fill = PatternFill(fill_type="solid", start_color="ADD8E6", end_color="ADD8E6")   # blue for low
    green_fill = PatternFill(fill_type="solid", start_color="90EE90", end_color="90EE90")   # green for ok
    red_fill   = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")   # light red for high
    
    # Process each cell in the Room sheet.
    for r in range(room_sheet.min_row, room_sheet.max_row + 1):
        for c in range(room_sheet.min_column, room_sheet.max_column + 1):
            room_cell = room_sheet.cell(row=r, column=c)
            # For header areas (before data), simply copy the value.
            if r < room_data_start_row or c < room_data_start_col:
                new_value = room_cell.value
            else:
                # Use the header row (assumed row 2) to identify the sensor ID.
                header_val = room_sheet.cell(row=2, column=c).value
                sensor_id = None
                if header_val and isinstance(header_val, str):
                    match = re.search(r"OC\d+", header_val)
                    if match:
                        sensor_id = match.group()
                
                if sensor_id:
                    # Get corresponding column indexes from Min and Max sheets.
                    min_col = min_mapping.get(sensor_id)
                    max_col = max_mapping.get(sensor_id)
                    room_value = room_sheet.cell(row=r, column=c).value
                    if min_col and max_col:
                        # Assuming the data rows are aligned across sheets.
                        min_value = min_sheet.cell(row=r, column=min_col).value
                        max_value = max_sheet.cell(row=r, column=max_col).value
                        
                        if all(isinstance(val, (int, float)) for val in [room_value, min_value, max_value]):
                            if room_value <= min_value:
                                diff = round(room_value - min_value, 2)
                                new_value = f"low: {diff}"
                            elif room_value >= max_value:
                                diff = round(room_value - max_value, 2)
                                new_value = f"high: {diff}"
                            else:
                                new_value = "ok"
                        else:
                            new_value = room_value
                    else:
                        # If sensor mapping is missing in one of the sheets, fall back to room value.
                        new_value = room_value
                else:
                    new_value = room_cell.value
            
            # Write the computed value to the Result sheet.
            result_cell = result_sheet.cell(row=r, column=c, value=new_value)
            if r >= room_data_start_row and c >= room_data_start_col and isinstance(new_value, str):
                if new_value.startswith("low:"):
                    result_cell.fill = blue_fill
                elif new_value.startswith("high:"):
                    result_cell.fill = red_fill
                elif new_value == "ok":
                    result_cell.fill = green_fill
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def main():
    st.title("Excel Temperature Checker")
    uploaded_file = st.file_uploader("Upload your Excel (.xlsx) file", type=["xlsx"])
    if uploaded_file:
        processed_file = process_excel(uploaded_file)
        if processed_file is not None:
            st.success("File processed successfully!")
            st.download_button(
                label="Download updated Excel file",
                data=processed_file,
                file_name=f"processed_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

if __name__ == "__main__":
    main()
